import openpyxl, io
from pathlib import Path
from PIL import Image
from .caption_api import auto_caption
from .settings import OUT_DIR

def _save_img(img, out_dir, idx):
    p = out_dir / f"img{idx}.png"
    img.ref.save(p); return p

def _save_chart(chart, out_dir, idx):
    buf = io.BytesIO(); chart._chart.render(buf)
    Image.open(buf).convert("RGB").save(out_dir / f"chart{idx}.png")
    return out_dir / f"chart{idx}.png"

def process_xlsx(path: Path, push_record):
    wb = openpyxl.load_workbook(path, data_only=False, read_only=True)
    doc_id = path.stem
    for ws in wb.worksheets:
        out_dir = OUT_DIR / doc_id / ws.title; out_dir.mkdir(parents=True, exist_ok=True)
        idx = 0
        for img in ws._images:
            p = _save_img(img, out_dir, idx)
            addr = f"{openpyxl.utils.get_column_letter(img.anchor._from.col+1)}{img.anchor._from.row+1}"
            cell_val = (ws[addr].value or "").strip()
            caption = cell_val or auto_caption(p)
            push_record(doc_id, f"{ws.title}_img{idx}", caption, p, raw_text=cell_val)
            idx += 1
        for c_i, chart in enumerate(ws._charts):
            p = _save_chart(chart, out_dir, c_i)
            chart_cap = chart.title.tx.rich.p[0].r[0].t if chart.title else ""
            caption = chart_cap or auto_caption(p)
            push_record(doc_id, f"{ws.title}_chart{c_i}", caption, p, raw_text=chart_cap)
