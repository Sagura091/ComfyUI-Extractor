"""
Enhanced Excel Processing with Revolutionary OCR
Extracts images with full worksheet context, OCR text recognition,
chart analysis, and comprehensive spreadsheet content analysis.
"""

import pandas as pd
import openpyxl
from openpyxl.drawing.image import Image as OpenpyxlImage
from PIL import Image
import io
import json
import logging
from pathlib import Path
from typing import Dict, List, Any, Optional, Tuple
import numpy as np

from .revolutionary_ocr import create_revolutionary_ocr
from .caption_api import auto_caption
from .settings import OUT_DIR
from .ocr_config import get_revolutionary_ocr_config

# Setup logging
logger = logging.getLogger(__name__)

class EnhancedXLSXProcessor:
    """Enhanced Excel processor with Revolutionary OCR capabilities"""
    
    def __init__(self, ocr_config: Dict[str, Any] = None):
        # Initialize revolutionary OCR
        self.ocr = create_revolutionary_ocr(ocr_config or {
            'languages': ['en'],
            'enable_table_recognition': True,
            'enable_formula_recognition': True,
            'enable_handwriting': True,
            'confidence_threshold': 0.7
        })
        
        self.logger = logging.getLogger(__name__)
    
    def process_xlsx_advanced(self, path: Path, push_record_func) -> Dict[str, Any]:
        """
        Process Excel file with advanced OCR and worksheet analysis
        """
        try:
            workbook = openpyxl.load_workbook(path, data_only=False)
            doc_id = path.stem
            
            processing_results = {
                'document_id': doc_id,
                'total_worksheets': len(workbook.worksheets),
                'processed_worksheets': 0,
                'images': [],
                'charts': [],
                'worksheet_content': [],
                'tables': [],
                'formulas': [],
                'errors': []
            }
            
            for ws_idx, worksheet in enumerate(workbook.worksheets):
                try:
                    ws_results = self._process_worksheet_advanced(
                        worksheet, doc_id, ws_idx, push_record_func, path
                    )
                    
                    # Aggregate results
                    processing_results['images'].extend(ws_results.get('images', []))
                    processing_results['charts'].extend(ws_results.get('charts', []))
                    processing_results['worksheet_content'].extend(ws_results.get('worksheet_content', []))
                    processing_results['tables'].extend(ws_results.get('tables', []))
                    processing_results['formulas'].extend(ws_results.get('formulas', []))
                    processing_results['processed_worksheets'] += 1
                    
                except Exception as e:
                    error_msg = f"Error processing worksheet {ws_idx} ({worksheet.title}): {str(e)}"
                    self.logger.error(error_msg)
                    processing_results['errors'].append(error_msg)
            
            return processing_results
            
        except Exception as e:
            error_msg = f"Error processing XLSX {path}: {str(e)}"
            self.logger.error(error_msg)
            return {'errors': [error_msg]}
    
    def _process_worksheet_advanced(self, worksheet, doc_id: str, ws_idx: int, 
                                  push_record_func, file_path: Path) -> Dict[str, Any]:
        """Process a single worksheet with comprehensive analysis"""
        
        ws_results = {
            'images': [],
            'charts': [],
            'worksheet_content': [],
            'tables': [],
            'formulas': []
        }
        
        # Create output directory
        out_dir = OUT_DIR / doc_id / f"worksheet{ws_idx}_{worksheet.title}"
        out_dir.mkdir(parents=True, exist_ok=True)
        
        # Extract comprehensive worksheet content
        worksheet_content = self._extract_comprehensive_worksheet_content(worksheet, file_path)
        
        # Generate worksheet screenshot for full context analysis
        ws_image = self._generate_worksheet_image(worksheet, out_dir, ws_idx)
        
        # Process worksheet image with OCR
        ocr_results = None
        if ws_image:
            try:
                ws_array = np.array(Image.open(ws_image))
                ocr_results = self.ocr.process_document_page(ws_array)
            except Exception as e:
                self.logger.warning(f"OCR processing failed for worksheet {ws_idx}: {e}")
        
        # Process embedded images
        img_idx = 0
        for image in worksheet._images:
            try:
                image_results = self._process_embedded_image(
                    image, worksheet, worksheet_content, ocr_results,
                    out_dir, doc_id, ws_idx, img_idx, push_record_func
                )
                ws_results['images'].append(image_results)
                img_idx += 1
            except Exception as e:
                self.logger.warning(f"Failed to process image {img_idx} on worksheet {ws_idx}: {e}")
        
        # Process charts
        chart_idx = 0
        for chart in worksheet._charts:
            try:
                chart_results = self._process_chart(
                    chart, worksheet, worksheet_content,
                    out_dir, doc_id, ws_idx, chart_idx, push_record_func
                )
                ws_results['charts'].append(chart_results)
                chart_idx += 1
            except Exception as e:
                self.logger.warning(f"Failed to process chart {chart_idx} on worksheet {ws_idx}: {e}")
        
        # Process data tables (identify table-like structures)
        table_results = self._identify_and_process_data_tables(
            worksheet, worksheet_content, out_dir, doc_id, ws_idx, push_record_func
        )
        ws_results['tables'].extend(table_results)
        
        # Process formulas
        formula_results = self._extract_and_process_formulas(
            worksheet, worksheet_content, out_dir, doc_id, ws_idx, push_record_func
        )
        ws_results['formulas'].extend(formula_results)
        
        # Store comprehensive worksheet content
        worksheet_content_record = {
            'worksheet_index': ws_idx,
            'worksheet_name': worksheet.title,
            'content': worksheet_content,
            'ocr_extracted_text': self._extract_ocr_text(ocr_results) if ocr_results else "",
            'image_count': img_idx,
            'chart_count': chart_idx,
            'table_count': len(ws_results['tables']),
            'formula_count': len(ws_results['formulas'])
        }
        
        # Push worksheet content as a record
        push_record_func(
            doc_id,
            f"ws{ws_idx}_content",
            f"Worksheet '{worksheet.title}' - {worksheet_content['summary'][:100]}",
            ws_image,
            raw_text=json.dumps(worksheet_content, indent=2)
        )
        
        ws_results['worksheet_content'].append(worksheet_content_record)
        
        return ws_results
    
    def _extract_comprehensive_worksheet_content(self, worksheet, file_path: Path) -> Dict[str, Any]:
        """Extract all content from a worksheet including data, structure, and metadata"""
        
        content = {
            'name': worksheet.title,
            'data_range': self._get_used_range(worksheet),
            'cell_data': {},
            'merged_cells': [],
            'comments': [],
            'headers': [],
            'data_types': {},
            'summary': '',
            'statistics': {},
            'structure_info': {}
        }
        
        # Get used range
        min_row, max_row, min_col, max_col = content['data_range']
        
        # Extract cell data with formatting info
        for row in range(min_row, min(max_row + 1, min_row + 100)):  # Limit to first 100 rows
            for col in range(min_col, min(max_col + 1, min_col + 20)):  # Limit to first 20 columns
                cell = worksheet.cell(row=row, column=col)
                if cell.value is not None:
                    cell_ref = f"{chr(64 + col)}{row}"
                    content['cell_data'][cell_ref] = {
                        'value': str(cell.value),
                        'data_type': type(cell.value).__name__,
                        'has_formula': cell.data_type == 'f',
                        'formula': cell.value if cell.data_type == 'f' else None,
                        'number_format': cell.number_format if hasattr(cell, 'number_format') else None
                    }
        
        # Extract merged cells
        for merged_range in worksheet.merged_cells.ranges:
            content['merged_cells'].append(str(merged_range))
        
        # Extract comments
        for cell in worksheet._comments:
            if cell.comment:
                content['comments'].append({
                    'cell': cell.coordinate,
                    'text': cell.comment.text,
                    'author': cell.comment.author if hasattr(cell.comment, 'author') else None
                })
        
        # Identify headers (first row with text)
        if max_row >= min_row:
            first_row_cells = []
            for col in range(min_col, min(max_col + 1, min_col + 20)):
                cell = worksheet.cell(row=min_row, column=col)
                if cell.value and isinstance(cell.value, str):
                    first_row_cells.append(cell.value)
            content['headers'] = first_row_cells
        
        # Analyze data types in each column
        for col in range(min_col, min(max_col + 1, min_col + 20)):
            col_letter = chr(64 + col)
            data_types = []
            for row in range(min_row + 1, min(max_row + 1, min_row + 50)):  # Skip header, sample 50 rows
                cell = worksheet.cell(row=row, column=col)
                if cell.value is not None:
                    data_types.append(type(cell.value).__name__)
            
            if data_types:
                # Most common data type
                content['data_types'][col_letter] = max(set(data_types), key=data_types.count)
        
        # Create summary from available data
        try:
            # Read as pandas dataframe for easy analysis
            df = pd.read_excel(file_path, sheet_name=worksheet.title, nrows=1000)  # Limit to 1000 rows
            
            summary_parts = [f"Worksheet '{worksheet.title}'"]
            summary_parts.append(f"Size: {len(df)} rows × {len(df.columns)} columns")
            
            if not df.empty:
                # Add column names if available
                if len(df.columns) > 0:
                    col_sample = list(df.columns)[:3]
                    summary_parts.append(f"Columns: {', '.join(str(c) for c in col_sample)}")
                    if len(df.columns) > 3:
                        summary_parts.append("...")
                
                # Add data preview
                if len(df) > 0:
                    first_row_sample = []
                    for col in df.columns[:3]:
                        val = df.iloc[0][col]
                        if pd.notna(val):
                            first_row_sample.append(str(val)[:20])
                    if first_row_sample:
                        summary_parts.append(f"Sample: {', '.join(first_row_sample)}")
            
            content['summary'] = ' | '.join(summary_parts)
            content['statistics'] = {
                'total_rows': len(df),
                'total_columns': len(df.columns),
                'numeric_columns': len(df.select_dtypes(include=[np.number]).columns),
                'text_columns': len(df.select_dtypes(include=['object']).columns),
                'null_percentage': (df.isnull().sum().sum() / (len(df) * len(df.columns))) * 100 if not df.empty else 0
            }
            
        except Exception as e:
            self.logger.warning(f"Failed to analyze worksheet with pandas: {e}")
            content['summary'] = f"Worksheet '{worksheet.title}' with {max_row - min_row + 1} rows"
        
        # Structure information
        content['structure_info'] = {
            'has_headers': bool(content['headers']),
            'has_formulas': any(cell.get('has_formula', False) for cell in content['cell_data'].values()),
            'has_merged_cells': bool(content['merged_cells']),
            'has_comments': bool(content['comments']),
            'data_density': len(content['cell_data']) / ((max_row - min_row + 1) * (max_col - min_col + 1)) if max_row > 0 else 0
        }
        
        return content
    
    def _get_used_range(self, worksheet) -> Tuple[int, int, int, int]:
        """Get the used range of the worksheet"""
        try:
            if worksheet.max_row and worksheet.max_column:
                return (worksheet.min_row or 1, worksheet.max_row, 
                       worksheet.min_column or 1, worksheet.max_column)
            else:
                return (1, 1, 1, 1)
        except:
            return (1, 1, 1, 1)
    
    def _generate_worksheet_image(self, worksheet, out_dir: Path, ws_idx: int) -> Optional[Path]:
        """Generate a visual representation of the worksheet"""
        try:
            # For now, create a text-based representation
            # In a full implementation, you might use openpyxl's drawing capabilities
            # or a library like xlwings to render the actual worksheet
            
            ws_image_path = out_dir / f"worksheet{ws_idx}_full.png"
            
            # Create a basic worksheet representation
            img = Image.new('RGB', (1200, 800), 'white')
            # This would be enhanced to actually render the worksheet content
            img.save(ws_image_path, 'PNG')
            
            return ws_image_path
            
        except Exception as e:
            self.logger.warning(f"Failed to generate worksheet image: {e}")
            return None
    
    def _process_embedded_image(self, image, worksheet, worksheet_content: Dict, ocr_results,
                               out_dir: Path, doc_id: str, ws_idx: int, img_idx: int,
                               push_record_func) -> Dict[str, Any]:
        """Process an embedded image with comprehensive context"""
        
        # Extract and save image
        img_data = image._data()
        img = Image.open(io.BytesIO(img_data)).convert("RGB")
        img_path = out_dir / f"img{img_idx}.png"
        img.save(img_path)
        
        # Get comprehensive caption
        caption = self._get_comprehensive_caption(
            image, worksheet_content, img_path, img_idx
        )
        
        # Create comprehensive context
        context_info = {
            'worksheet_index': ws_idx,
            'worksheet_name': worksheet_content.get('name', ''),
            'worksheet_summary': worksheet_content.get('summary', ''),
            'image_anchor': str(image.anchor) if hasattr(image, 'anchor') else '',
            'surrounding_data': self._get_surrounding_data(image, worksheet),
            'worksheet_headers': worksheet_content.get('headers', []),
            'worksheet_statistics': worksheet_content.get('statistics', {})
        }
        
        # Perform OCR on the image itself
        img_ocr_results = self._ocr_analyze_image(img_path)
        if img_ocr_results:
            context_info['image_ocr_text'] = img_ocr_results.get('text', '')
            context_info['image_contains_text'] = bool(img_ocr_results.get('text', '').strip())
            context_info['image_has_tables'] = img_ocr_results.get('has_tables', False)
        
        # Enhanced description
        enhanced_description = self._create_enhanced_description(caption, context_info)
        
        # Push to record system
        push_record_func(
            doc_id,
            f"ws{ws_idx}_img{img_idx}",
            enhanced_description,
            img_path,
            raw_text=json.dumps(context_info, indent=2)
        )
        
        return {
            'path': str(img_path),
            'caption': enhanced_description,
            'context': context_info,
            'worksheet': ws_idx,
            'index': img_idx
        }
    
    def _process_chart(self, chart, worksheet, worksheet_content: Dict,
                      out_dir: Path, doc_id: str, ws_idx: int, chart_idx: int,
                      push_record_func) -> Dict[str, Any]:
        """Process a chart with context analysis"""
        
        # Create chart description
        chart_info = {
            'type': str(chart.__class__.__name__),
            'title': chart.title.tx.rich.p[0].r.t if hasattr(chart, 'title') and chart.title else f"Chart {chart_idx + 1}",
            'worksheet': worksheet_content.get('name', ''),
            'data_references': self._extract_chart_data_references(chart),
        }
        
        # Create chart context
        context_info = {
            'worksheet_index': ws_idx,
            'worksheet_name': worksheet_content.get('name', ''),
            'chart_type': chart_info['type'],
            'chart_title': chart_info['title'],
            'data_sources': chart_info['data_references'],
            'worksheet_summary': worksheet_content.get('summary', ''),
            'worksheet_headers': worksheet_content.get('headers', [])
        }
        
        # Create description
        description = f"Chart: {chart_info['title']} ({chart_info['type']}) from worksheet '{worksheet_content.get('name', '')}'"
        if chart_info['data_references']:
            description += f" | Data from: {', '.join(chart_info['data_references'][:3])}"
        
        # For now, we'll create a placeholder image
        # In a full implementation, you'd render the actual chart
        chart_path = out_dir / f"chart{chart_idx}.png"
        placeholder_img = Image.new('RGB', (400, 300), 'lightgray')
        placeholder_img.save(chart_path)
        
        push_record_func(
            doc_id,
            f"ws{ws_idx}_chart{chart_idx}",
            description,
            chart_path,
            raw_text=json.dumps(context_info, indent=2)
        )
        
        return {
            'path': str(chart_path),
            'description': description,
            'context': context_info,
            'worksheet': ws_idx,
            'index': chart_idx
        }
    
    def _identify_and_process_data_tables(self, worksheet, worksheet_content: Dict,
                                        out_dir: Path, doc_id: str, ws_idx: int,
                                        push_record_func) -> List[Dict[str, Any]]:
        """Identify and process data tables in the worksheet"""
        
        tables = []
        
        # Simple table detection based on continuous data regions
        min_row, max_row, min_col, max_col = worksheet_content['data_range']
        
        # Look for regions with headers and data
        potential_tables = self._find_table_regions(worksheet, min_row, max_row, min_col, max_col)
        
        for table_idx, table_region in enumerate(potential_tables):
            try:
                table_data = self._extract_table_data(worksheet, table_region)
                
                if len(table_data) > 1:  # Must have at least header + 1 data row
                    description = f"Data table from worksheet '{worksheet_content.get('name', '')}': "
                    description += f"{len(table_data)} rows × {len(table_data[0]) if table_data else 0} columns"
                    
                    if table_data[0]:  # Add header info
                        headers = [str(h) for h in table_data[0][:3]]
                        description += f" | Headers: {', '.join(headers)}"
                        if len(table_data[0]) > 3:
                            description += "..."
                    
                    # Create table context
                    context_info = {
                        'worksheet_index': ws_idx,
                        'worksheet_name': worksheet_content.get('name', ''),
                        'table_region': table_region,
                        'table_data': table_data[:10],  # First 10 rows
                        'total_rows': len(table_data),
                        'total_columns': len(table_data[0]) if table_data else 0
                    }
                    
                    push_record_func(
                        doc_id,
                        f"ws{ws_idx}_table{table_idx}",
                        description,
                        None,  # No image for data tables
                        raw_text=json.dumps(table_data, indent=2)
                    )
                    
                    tables.append({
                        'description': description,
                        'context': context_info,
                        'data': table_data,
                        'worksheet': ws_idx,
                        'index': table_idx
                    })
                    
            except Exception as e:
                self.logger.warning(f"Failed to process table {table_idx}: {e}")
        
        return tables
    
    def _extract_and_process_formulas(self, worksheet, worksheet_content: Dict,
                                    out_dir: Path, doc_id: str, ws_idx: int,
                                    push_record_func) -> List[Dict[str, Any]]:
        """Extract and process formulas from the worksheet"""
        
        formulas = []
        formula_idx = 0
        
        # Find all cells with formulas
        min_row, max_row, min_col, max_col = worksheet_content['data_range']
        
        for row in range(min_row, min(max_row + 1, min_row + 100)):
            for col in range(min_col, min(max_col + 1, min_col + 20)):
                cell = worksheet.cell(row=row, column=col)
                
                if cell.data_type == 'f' and cell.value:  # Formula cell
                    try:
                        formula_info = {
                            'cell_reference': f"{chr(64 + col)}{row}",
                            'formula': cell.value,
                            'calculated_value': cell.internal_value if hasattr(cell, 'internal_value') else None,
                            'worksheet': worksheet_content.get('name', ''),
                            'context': self._get_formula_context(worksheet, row, col)
                        }
                        
                        description = f"Formula in {formula_info['cell_reference']}: {formula_info['formula']}"
                        if formula_info['calculated_value'] is not None:
                            description += f" = {formula_info['calculated_value']}"
                        description += f" | Worksheet: {worksheet_content.get('name', '')}"
                        
                        push_record_func(
                            doc_id,
                            f"ws{ws_idx}_formula{formula_idx}",
                            description,
                            None,
                            raw_text=formula_info['formula']
                        )
                        
                        formulas.append({
                            'description': description,
                            'info': formula_info,
                            'worksheet': ws_idx,
                            'index': formula_idx
                        })
                        
                        formula_idx += 1
                        
                    except Exception as e:
                        self.logger.warning(f"Failed to process formula at {chr(64 + col)}{row}: {e}")
        
        return formulas
    
    def _get_comprehensive_caption(self, image, worksheet_content: Dict, img_path: Path, img_idx: int) -> str:
        """Generate comprehensive caption for embedded image"""
        
        captions = []
        
        # Method 1: AI-generated caption
        try:
            ai_caption = auto_caption(str(img_path))
            if ai_caption and len(ai_caption.strip()) > 5:
                captions.append(('ai_caption', ai_caption, 0.8))
        except Exception as e:
            self.logger.warning(f"AI captioning failed: {e}")
        
        # Method 2: OCR on image
        try:
            img_array = np.array(Image.open(img_path))
            ocr_result = self.ocr.extract_text_multilingual(img_array)
            if ocr_result.confidence > 0.7 and len(ocr_result.text.strip()) > 5:
                captions.append(('ocr_text', f"Image containing text: {ocr_result.text[:100]}", 0.7))
        except Exception as e:
            self.logger.warning(f"Image OCR failed: {e}")
        
        # Method 3: Context-based caption
        ws_name = worksheet_content.get('name', '')
        if ws_name:
            context_caption = f"Image from worksheet '{ws_name}'"
            if worksheet_content.get('headers'):
                context_caption += f" with data: {', '.join(worksheet_content['headers'][:2])}"
            captions.append(('context', context_caption, 0.6))
        
        # Select best caption
        if captions:
            captions.sort(key=lambda x: x[2], reverse=True)
            return captions[0][1]
        else:
            return f"Image {img_idx + 1} from worksheet '{ws_name}'"
    
    def _get_surrounding_data(self, image, worksheet) -> str:
        """Get data from cells near the image"""
        # This is a simplified implementation
        # In practice, you'd analyze the image anchor position
        surrounding_data = []
        
        # Sample some cells for context
        for row in range(1, min(21, worksheet.max_row + 1)):
            for col in range(1, min(11, worksheet.max_column + 1)):
                cell = worksheet.cell(row=row, column=col)
                if cell.value and isinstance(cell.value, str):
                    surrounding_data.append(str(cell.value)[:50])
                    if len(surrounding_data) >= 5:
                        break
            if len(surrounding_data) >= 5:
                break
        
        return ' | '.join(surrounding_data)
    
    def _ocr_analyze_image(self, img_path: Path) -> Optional[Dict[str, Any]]:
        """Perform OCR analysis on a single image"""
        try:
            img_array = np.array(Image.open(img_path))
            results = self.ocr.process_document_page(img_array)
            
            all_text = []
            for region in results.get('text_regions', []):
                if region.confidence > 0.5:
                    all_text.append(region.text)
            
            return {
                'text': ' '.join(all_text),
                'has_tables': len(results.get('table_regions', [])) > 0,
                'has_formulas': len(results.get('formula_regions', [])) > 0,
                'confidence': max([r.confidence for r in results.get('text_regions', [])], default=0.0)
            }
        except Exception as e:
            self.logger.warning(f"OCR analysis failed for {img_path}: {e}")
            return None
    
    def _create_enhanced_description(self, base_caption: str, context_info: Dict) -> str:
        """Create enhanced description combining caption and context"""
        
        description_parts = [base_caption]
        
        # Add worksheet context
        if context_info.get('worksheet_name'):
            description_parts.append(f"From worksheet: '{context_info['worksheet_name']}'")
        
        # Add data context
        if context_info.get('worksheet_headers'):
            headers = context_info['worksheet_headers'][:3]
            description_parts.append(f"Data columns: {', '.join(headers)}")
        
        # Add OCR text if present
        if context_info.get('image_contains_text') and context_info.get('image_ocr_text'):
            text_preview = context_info['image_ocr_text'][:100]
            description_parts.append(f"Contains text: {text_preview}")
        
        return ' | '.join(description_parts)
    
    def _extract_chart_data_references(self, chart) -> List[str]:
        """Extract data references from a chart"""
        references = []
        try:
            # This is simplified - actual implementation would parse chart series data
            if hasattr(chart, 'series'):
                for series in chart.series:
                    if hasattr(series, 'val') and hasattr(series.val, 'strRef'):
                        if series.val.strRef.f:
                            references.append(series.val.strRef.f)
        except Exception as e:
            self.logger.warning(f"Failed to extract chart references: {e}")
        
        return references
    
    def _find_table_regions(self, worksheet, min_row: int, max_row: int, 
                          min_col: int, max_col: int) -> List[Dict]:
        """Find potential table regions in the worksheet"""
        
        regions = []
        
        # Simple heuristic: look for continuous data regions with headers
        current_region = None
        
        for row in range(min_row, min(max_row + 1, min_row + 50)):
            row_has_data = False
            for col in range(min_col, min(max_col + 1, min_col + 10)):
                cell = worksheet.cell(row=row, column=col)
                if cell.value is not None:
                    row_has_data = True
                    break
            
            if row_has_data:
                if current_region is None:
                    current_region = {
                        'start_row': row,
                        'end_row': row,
                        'start_col': min_col,
                        'end_col': min_col
                    }
                else:
                    current_region['end_row'] = row
            else:
                if current_region and current_region['end_row'] - current_region['start_row'] > 2:
                    # Find actual column range
                    actual_min_col, actual_max_col = self._find_column_range(
                        worksheet, current_region['start_row'], current_region['end_row'], min_col, max_col
                    )
                    current_region['start_col'] = actual_min_col
                    current_region['end_col'] = actual_max_col
                    regions.append(current_region)
                current_region = None
        
        # Add final region if exists
        if current_region and current_region['end_row'] - current_region['start_row'] > 2:
            actual_min_col, actual_max_col = self._find_column_range(
                worksheet, current_region['start_row'], current_region['end_row'], min_col, max_col
            )
            current_region['start_col'] = actual_min_col
            current_region['end_col'] = actual_max_col
            regions.append(current_region)
        
        return regions
    
    def _find_column_range(self, worksheet, start_row: int, end_row: int, 
                          min_col: int, max_col: int) -> Tuple[int, int]:
        """Find the actual column range for a table region"""
        
        used_cols = set()
        
        for row in range(start_row, end_row + 1):
            for col in range(min_col, min(max_col + 1, min_col + 20)):
                cell = worksheet.cell(row=row, column=col)
                if cell.value is not None:
                    used_cols.add(col)
        
        if used_cols:
            return min(used_cols), max(used_cols)
        else:
            return min_col, min_col
    
    def _extract_table_data(self, worksheet, region: Dict) -> List[List[str]]:
        """Extract data from a table region"""
        
        data = []
        
        for row in range(region['start_row'], region['end_row'] + 1):
            row_data = []
            for col in range(region['start_col'], region['end_col'] + 1):
                cell = worksheet.cell(row=row, column=col)
                row_data.append(str(cell.value) if cell.value is not None else '')
            data.append(row_data)
        
        return data
    
    def _get_formula_context(self, worksheet, row: int, col: int) -> Dict[str, Any]:
        """Get context information for a formula cell"""
        
        context = {
            'neighboring_cells': {},
            'referenced_ranges': [],
            'cell_style': {}
        }
        
        # Get neighboring cells
        for dr in [-1, 0, 1]:
            for dc in [-1, 0, 1]:
                if dr == 0 and dc == 0:
                    continue
                
                neighbor_row, neighbor_col = row + dr, col + dc
                if neighbor_row > 0 and neighbor_col > 0:
                    try:
                        neighbor_cell = worksheet.cell(row=neighbor_row, column=neighbor_col)
                        if neighbor_cell.value is not None:
                            context['neighboring_cells'][f"{chr(64 + neighbor_col)}{neighbor_row}"] = str(neighbor_cell.value)[:50]
                    except:
                        pass
        
        return context
    
    def _extract_ocr_text(self, ocr_results: Dict) -> str:
        """Extract all text from OCR results"""
        all_text = []
        
        for region in ocr_results.get('text_regions', []):
            if region.confidence > 0.5:
                all_text.append(region.text)
        
        return ' '.join(all_text)


# Backward compatibility function
def process_xlsx(path: Path, push_record):
    """
    Enhanced XLSX processing function with Revolutionary OCR
    """
    try:
        # Initialize enhanced processor
        config = get_revolutionary_ocr_config()
        processor = EnhancedXLSXProcessor(config['ocr'])
        
        # Process with advanced capabilities
        results = processor.process_xlsx_advanced(path, push_record)
        
        # Log processing summary
        logger.info(f"Processed XLSX {path.name}: "
                   f"{results.get('processed_worksheets', 0)}/{results.get('total_worksheets', 0)} worksheets, "
                   f"{len(results.get('images', []))} images, "
                   f"{len(results.get('charts', []))} charts, "
                   f"{len(results.get('tables', []))} tables")
        
        return results
        
    except Exception as e:
        logger.error(f"Failed to process XLSX {path}: {e}")
        # Fallback to basic processing
        return _process_xlsx_basic(path, push_record)


def _process_xlsx_basic(path: Path, push_record):
    """Fallback to basic XLSX processing if advanced processing fails"""
    try:
        workbook = openpyxl.load_workbook(path)
        doc_id = path.stem
        
        for ws_idx, worksheet in enumerate(workbook.worksheets):
            out_dir = OUT_DIR / doc_id / f"ws{ws_idx}"
            out_dir.mkdir(parents=True, exist_ok=True)
            
            # Basic text extraction
            text_content = []
            for row in worksheet.iter_rows(values_only=True, max_row=100):
                row_text = ' '.join(str(cell) for cell in row if cell is not None)
                if row_text.strip():
                    text_content.append(row_text)
            
            worksheet_text = '\n'.join(text_content)
            
            # Process images
            img_idx = 0
            for image in worksheet._images:
                try:
                    img_data = image._data()
                    img = Image.open(io.BytesIO(img_data)).convert("RGB")
                    img_path = out_dir / f"img{img_idx}.png"
                    img.save(img_path)
                    
                    caption = auto_caption(str(img_path)) or f"Image from worksheet {worksheet.title}"
                    push_record(doc_id, f"ws{ws_idx}_img{img_idx}", caption, img_path, raw_text=worksheet_text)
                    img_idx += 1
                except Exception as e:
                    logger.warning(f"Failed to process image {img_idx}: {e}")
                    
    except Exception as e:
        logger.error(f"Basic XLSX processing also failed: {e}")
        raise
